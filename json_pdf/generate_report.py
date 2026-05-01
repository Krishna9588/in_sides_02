"""
generate_report.py
==================
Reusable script: converts any db_document.json → professional Excel + DOCX report.

Usage:
    python generate_report.py <path/to/db_document.json>
    python generate_report.py db_document.json          # defaults to current dir

Outputs:
    output/<company_name>/
        ├── <company_name>_report.xlsx
        └── <company_name>_report.docx
"""

import sys, os, re, json, io, textwrap, subprocess, tempfile
from pathlib import Path
from datetime import datetime

import requests
from PIL import Image as PILImage

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════════

LOGO_DEV_KEY   = "pk_DOneOcGwSAau_ztcDSzJYw"
LOGO_SIZE_PX   = 80          # px we'll request from logo.dev
FONT_NAME      = "Calibri"

# Brand palette
C_DARK         = "0D1F3C"    # deep navy
C_ACCENT       = "1A6BCC"    # corporate blue
C_ACCENT_LIGHT = "D6E8FF"    # pale blue fill
C_GREEN        = "1A7A4A"    # success green
C_GREEN_LIGHT  = "D4EDDA"
C_AMBER        = "B76E00"    # warning amber
C_AMBER_LIGHT  = "FFF3CD"
C_RED          = "B71C1C"    # danger red
C_RED_LIGHT    = "FDECEA"
C_GREY_BG      = "F5F7FA"    # alternating row
C_GREY_LINE    = "DEE2E8"    # border colour
C_WHITE        = "FFFFFF"
C_TEXT_DARK    = "1A1A2E"
C_TEXT_MUTED   = "6B7280"

# Severity → colour mapping (Excel and DOCX)
SEV_PALETTE = {
    "critical": (C_RED,      C_RED_LIGHT),
    "high":     (C_AMBER,    C_AMBER_LIGHT),
    "medium":   ("7B5800",   "FFFBEA"),
    "low":      (C_GREEN,    C_GREEN_LIGHT),
    "p0":       (C_RED,      C_RED_LIGHT),
    "p1":       (C_AMBER,    C_AMBER_LIGHT),
    "p2":       ("7B5800",   "FFFBEA"),
    "p3":       (C_GREEN,    C_GREEN_LIGHT),
}

# ══════════════════════════════════════════════════════════════════════════════
# UTILITY — safe value extraction, cleaning, formatting
# ══════════════════════════════════════════════════════════════════════════════

_NULL_VALS = {None, "", "null", "none", "n/a", "na", "—", "unable to verify",
              "not available", "not found", "undefined"}

def is_empty(v):
    """Return True when a value carries no useful information."""
    if v is None:
        return True
    if isinstance(v, str) and v.strip().lower() in _NULL_VALS:
        return True
    if isinstance(v, (list, dict)) and len(v) == 0:
        return True
    return False

def safe(v, fallback="—"):
    """Scalar value → clean string, empty → fallback."""
    if is_empty(v):
        return fallback
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, float):
        return f"{v:,.4g}"
    if isinstance(v, list):
        parts = [safe(i) for i in v if not is_empty(i)]
        return ", ".join(parts) if parts else fallback
    return str(v).strip()

def clean_text(v):
    """Remove internal JSON artefacts from long text fields."""
    s = safe(v)
    if s == "—":
        return s
    # Remove leftover raw URLs that are wrapped in markdown [text](url)
    s = re.sub(r'\[([^\]]+)\]\(https?://[^\)]+\)', r'\1', s)
    # Collapse 3+ newlines
    s = re.sub(r'\n{3,}', '\n\n', s)
    # Remove leading/trailing whitespace per line
    lines = [l.rstrip() for l in s.split('\n')]
    return '\n'.join(lines).strip()

def is_url(v):
    return isinstance(v, str) and re.match(r'https?://', v.strip())

def fmt_date(v, fallback="—"):
    if is_empty(v):
        return fallback
    s = str(v)
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f",
                "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%SZ"):
        try:
            return datetime.strptime(s[:26], fmt).strftime("%-d %b %Y")
        except Exception:
            pass
    return s[:10]

def fmt_num(v, suffix=""):
    if is_empty(v):
        return "—"
    try:
        return f"{int(float(str(v))):,}{suffix}"
    except Exception:
        return safe(v)

def star_bar(rating, max_stars=5):
    """Return a star-bar string, e.g. ★★★★☆ 4.3"""
    try:
        r = float(rating)
        full = int(r)
        half = 1 if (r - full) >= 0.5 else 0
        empty = max_stars - full - half
        bar = "★" * full + ("½" if half else "") + "☆" * empty
        return f"{bar}  {r:.1f}"
    except Exception:
        return safe(rating)

def clean_domain(domain_str):
    """https://univest.in/ → univest.in"""
    d = re.sub(r'^https?://', '', str(domain_str or "")).rstrip('/')
    return d

def slug(name):
    return re.sub(r'[^\w]+', '_', str(name).strip()).strip('_')

# ══════════════════════════════════════════════════════════════════════════════
# LOGO FETCH
# ══════════════════════════════════════════════════════════════════════════════

def fetch_logo_bytes(domain, company_name):
    """Try logo.dev first (domain, then name); fall back to initials PNG."""
    domain_clean = clean_domain(domain)
    for query in [domain_clean, re.sub(r'\s+', '', company_name.lower())]:
        try:
            url = f"https://img.logo.dev/{query}?token={LOGO_DEV_KEY}&size={LOGO_SIZE_PX}&format=png"
            r = requests.get(url, timeout=6, headers={"User-Agent": "Mozilla/5.0"})
            if r.status_code == 200 and r.headers.get("content-type", "").startswith("image"):
                return r.content, "png"
        except Exception:
            pass
    # Fallback: generate an initials badge with Pillow
    return _make_initials_logo(company_name), "png"

def _make_initials_logo(name):
    """Create a coloured square with company initials using Pillow."""
    from PIL import ImageDraw, ImageFont
    size = LOGO_SIZE_PX * 2          # render at 2x, downscale
    img  = PILImage.new("RGB", (size, size), color=(26, 107, 204))
    draw = ImageDraw.Draw(img)
    initials = "".join(w[0].upper() for w in name.split()[:2])
    # Simple font sizing
    font_size = size // 2
    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", font_size)
    except Exception:
        font = ImageFont.load_default()
    bbox = draw.textbbox((0, 0), initials, font=font)
    x = (size - (bbox[2] - bbox[0])) // 2 - bbox[0]
    y = (size - (bbox[3] - bbox[1])) // 2 - bbox[1]
    draw.text((x, y), initials, fill=(255, 255, 255), font=font)
    buf = io.BytesIO()
    img = img.resize((LOGO_SIZE_PX, LOGO_SIZE_PX), PILImage.LANCZOS)
    img.save(buf, format="PNG")
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# DATA EXTRACTOR — normalises every section from JSON
# ══════════════════════════════════════════════════════════════════════════════

class ReportData:
    """Parses db_document.json and exposes clean, typed attributes."""

    def __init__(self, path):
        with open(path, encoding="utf-8") as f:
            raw = json.load(f)

        self.project_name  = safe(raw.get("project_name"))
        self.domain        = safe(raw.get("domain"))
        self.ingestion_date = fmt_date(raw.get("ingestion_date"))

        src = raw.get("data_sources", {})

        # ── Company Profile ──────────────────────────────────────────────────
        cp = src.get("company_profile", {}).get("data", {})
        self.company_name       = safe(cp.get("company_name"), self.project_name)
        self.company_domain     = safe(cp.get("domain"), self.domain)
        self.playstore_link     = safe(cp.get("playstore_link"))
        self.appstore_link      = safe(cp.get("appstore_link"))
        self.youtube_channel    = safe(cp.get("youtube_official_channel"))
        self.linkedin_page      = safe(cp.get("linkedin_company_page"))
        self.year_founded       = safe(cp.get("year_founded"))
        self.hq_location        = safe(cp.get("exact_hq_location"))
        self.locations          = self._flatten_list(cp.get("locations_operating_in", []))
        self.industry           = clean_text(cp.get("industry_and_segment"))
        self.platforms          = safe(cp.get("available_platforms"))
        self.employee_count     = safe(cp.get("employee_count"))
        self.funding_raised     = safe(cp.get("funding_raised"))
        self.funding_stage      = safe(cp.get("funding_stage"))
        self.no_of_users        = safe(cp.get("no_of_users"))
        self.annual_revenue     = safe(cp.get("annual_revenue"))
        self.key_positioning    = clean_text(cp.get("key_positioning"))
        self.revenue_model      = safe(cp.get("revenue_model"))
        self.pricing_tiers      = self._flatten_list(cp.get("pricing_tiers", []))
        self.target_segments    = self._flatten_list(cp.get("target_customer_segments", []))
        self.tech_stack         = self._flatten_list(cp.get("tech_stack_highlights", []))
        self.milestones         = self._flatten_list(cp.get("milestones", []))
        self.new_features       = self._flatten_list(cp.get("new_features_launched", []))
        self.founders           = self._flatten_list(cp.get("names_of_founders", []))
        self.csuite             = self._flatten_list(cp.get("c-suite_officer", cp.get("c_suite_officer", [])))
        self.competitors        = self._parse_structured_list(cp.get("competitors", []))
        self.partnerships       = self._parse_structured_list(cp.get("recent_partnerships_and_integrations", []))
        self.strategic_moves    = self._parse_structured_list(cp.get("strategic_moves", []))
        self.differentiators    = self._parse_structured_list(cp.get("differentiators", []))
        self.user_complaints    = self._parse_structured_list(cp.get("user_complaints", []))
        self.regulatory         = self._parse_structured_list(cp.get("regulatory_and_legal_issues", []))
        self.current_problems   = self._parse_structured_list(cp.get("current_problems_struggling_with", []))
        self.other_details      = self._flatten_list(cp.get("other_crucial_details", []))
        self.market_sentiment   = cp.get("market_sentiment") or {}

        # ── Play Store ───────────────────────────────────────────────────────
        ps_raw = src.get("play_store", {})
        ps_meta = ps_raw.get("extracted_data", {}).get("metadata", {})
        ps_ra   = ps_raw.get("extracted_data", {}).get("review_analysis", {})
        ps_rev  = ps_raw.get("extracted_data", {}).get("reviews", [])
        self.ps = {
            "title":          safe(ps_meta.get("title")),
            "developer":      safe(ps_meta.get("developer")),
            "score":          safe(ps_meta.get("score")),
            "ratings":        fmt_num(ps_meta.get("ratings")),
            "reviews":        fmt_num(ps_meta.get("reviews")),
            "installs":       safe(ps_meta.get("installs")),
            "genre":          safe(ps_meta.get("genre")),
            "released":       safe(ps_meta.get("released")),
            "version":        safe(ps_meta.get("version")),
            "content_rating": safe(ps_meta.get("content_rating")),
            "free":           "Free" if ps_meta.get("free") else "Paid",
            "url":            safe(ps_meta.get("url")),
            "privacy_policy": safe(ps_meta.get("developer_privacy_policy")),
            "min_android":    safe(ps_meta.get("min_android_version")),
            "total_reviews":  fmt_num(ps_ra.get("total_reviews")),
            "avg_rating":     safe(ps_ra.get("average_rating")),
            "latest_review":  fmt_date(ps_ra.get("latest_review_date")),
            "oldest_review":  fmt_date(ps_ra.get("oldest_review_date")),
            "rating_dist":    ps_ra.get("rating_distribution", {}),
            "reviews_list":   self._parse_ps_reviews(ps_rev),
        }

        # ── App Store ────────────────────────────────────────────────────────
        as_raw  = src.get("app_store", {})
        as_meta = as_raw.get("extracted_data", {}).get("metadata", {})
        self.app = {
            "title":         safe(as_meta.get("trackName")),
            "developer":     safe(as_meta.get("artistName")),
            "score":         safe(as_meta.get("averageUserRating")),
            "ratings":       fmt_num(as_meta.get("userRatingCount")),
            "version":       safe(as_meta.get("version")),
            "min_ios":       safe(as_meta.get("minimumOsVersion")),
            "genre":         safe(as_meta.get("primaryGenreName")),
            "released":      fmt_date(as_meta.get("releaseDate")),
            "price":         safe(as_meta.get("formattedPrice")),
            "url":           safe(as_meta.get("trackViewUrl")),
            "content_rating":safe(as_meta.get("contentAdvisoryRating")),
        }

        # ── Transcripts ──────────────────────────────────────────────────────
        tr = src.get("internal_transcripts", {})
        self.transcript = {
            "source_file":   safe(tr.get("source_file")),
            "total_signals": safe(tr.get("total_signals")),
            "classifier":    safe(tr.get("classifier_used")),
            "meeting_type":  safe((tr.get("metadata") or {}).get("meeting_type")),
            "processed_at":  fmt_date((tr.get("metadata") or {}).get("processed_at")),
            "signals":       self._parse_signals(tr.get("signals", [])),
        }

        # ── Reddit ───────────────────────────────────────────────────────────
        reddit_raw = src.get("reddit", {})
        self.reddit_posts = []
        for block in reddit_raw.values():
            for post in block.get("posts", []):
                if not is_empty(post.get("title")):
                    self.reddit_posts.append({
                        "title":        clean_text(post.get("title")),
                        "subreddit":    safe(post.get("subreddit")),
                        "author":       safe(post.get("author")),
                        "score":        fmt_num(post.get("score")),
                        "comments":     fmt_num(post.get("num_comments")),
                        "url":          safe(post.get("url")),
                        "selftext":     clean_text(post.get("selftext")),
                    })

        # ── YouTube ──────────────────────────────────────────────────────────
        yt_raw = src.get("youtube", {})
        self.youtube_videos = []
        for block in yt_raw.values():
            for vid in (block if isinstance(block, list) else []):
                if not is_empty(vid.get("title")):
                    self.youtube_videos.append({
                        "title":       clean_text(vid.get("title")),
                        "url":         safe(vid.get("url")),
                        "video_id":    safe(vid.get("video_id")),
                        "description": clean_text(vid.get("description")),
                        "views":       fmt_num(vid.get("view_count")),
                        "likes":       fmt_num(vid.get("like_count")),
                        "published":   fmt_date(vid.get("published_at")),
                        "scraped_at":  fmt_date(vid.get("scraped_at")),
                    })

        # ── Agent 2 — Problems ───────────────────────────────────────────────
        ag2 = raw.get("agent2_output", {})
        self.problems = []
        for p in ag2.get("problems", []):
            self.problems.append({
                "id":               safe(p.get("problem_id")),
                "problem":          clean_text(p.get("problem")),
                "severity":         safe(p.get("severity")),
                "frequency":        safe(p.get("frequency")),
                "category":         safe(p.get("category")),
                "user_type":        safe(p.get("user_type")),
                "sources":          self._flatten_list(p.get("source_mix", [])),
                "competitor_issue": safe(p.get("competitor_has_same_issue")),
                "evidence":         [clean_text(e) for e in (p.get("evidence") or []) if not is_empty(e)],
            })
        self.total_problems     = safe(ag2.get("total_problems"))
        self.top_categories     = self._flatten_list(ag2.get("top_categories", []))
        self.high_severity_count= safe(ag2.get("high_severity_count"))

        # ── Agent 3 — Insights ───────────────────────────────────────────────
        ag3 = raw.get("agent3_output", {})
        self.insights = []
        for i in ag3.get("insights", []):
            self.insights.append({
                "id":              safe(i.get("insight_id")),
                "insight":         clean_text(i.get("insight")),
                "priority":        safe(i.get("priority")),
                "confidence":      safe(i.get("confidence")),
                "theme":           safe(i.get("theme")),
                "root_cause":      clean_text(i.get("root_cause")),
                "evidence":        clean_text(i.get("evidence_summary")),
                "competitor_gap":  clean_text(i.get("competitor_gap")),
                "opportunity":     clean_text(i.get("opportunity_size")),
                "implication":     clean_text(i.get("implication")),
                "support_ids":     self._flatten_list(i.get("supporting_problem_ids", [])),
            })
        self.total_insights     = safe(ag3.get("total_insights"))
        self.critical_count     = safe(ag3.get("critical_count"))
        self.dominant_theme     = safe(ag3.get("dominant_theme"))
        self.strategic_risk     = clean_text(ag3.get("key_strategic_risk"))
        self.biggest_opp        = clean_text(ag3.get("biggest_opportunity"))

        # ── Agent 4 — Briefs ─────────────────────────────────────────────────
        ag4 = raw.get("agent4_output", {})
        self.briefs = []
        for b in ag4.get("briefs", []):
            self.briefs.append({
                "id":           safe(b.get("brief_id")),
                "feature":      safe(b.get("feature_name")),
                "priority":     safe(b.get("priority")),
                "effort":       safe(b.get("effort")),
                "insight_ref":  safe(b.get("addresses_insight")),
                "problem":      clean_text(b.get("problem")),
                "why_now":      clean_text(b.get("why_now")),
                "solution":     clean_text(b.get("solution")),
                "impact":       clean_text(b.get("expected_impact")),
                "metric":       clean_text(b.get("success_metric")),
                "user_flow":    [clean_text(s) for s in (b.get("user_flow") or []) if not is_empty(s)],
            })
        self.total_briefs       = safe(ag4.get("total_briefs"))
        self.sprint_focus       = clean_text(ag4.get("recommended_sprint_focus"))

    # ── helpers ──────────────────────────────────────────────────────────────

    def _flatten_list(self, lst):
        out = []
        for item in (lst or []):
            if is_empty(item):
                continue
            if isinstance(item, str):
                out.append(item.strip())
            elif isinstance(item, dict):
                # Pick the most descriptive text key
                for key in ("description", "feature", "move", "issue", "detail", "note", "text"):
                    if not is_empty(item.get(key)):
                        out.append(clean_text(item[key]))
                        break
                else:
                    # Fall back to joining all non-url values
                    parts = [str(v) for v in item.values()
                             if not is_empty(v) and not is_url(str(v))]
                    if parts:
                        out.append(" | ".join(parts))
            else:
                out.append(str(item))
        return out

    def _parse_structured_list(self, lst):
        """Return list of dicts, skip entries where all useful fields are empty."""
        out = []
        for item in (lst or []):
            if is_empty(item):
                continue
            if isinstance(item, str):
                out.append({"value": item})
            elif isinstance(item, dict):
                cleaned = {k: v for k, v in item.items()
                           if not is_empty(v) and k not in ("effect",)}
                if cleaned:
                    out.append(cleaned)
        return out

    def _parse_ps_reviews(self, lst):
        out = []
        for rev in (lst or []):
            content = clean_text(rev.get("content") or rev.get("text"))
            if is_empty(content):
                continue
            out.append({
                "author":  safe(rev.get("author") or rev.get("userName")),
                "rating":  safe(rev.get("rating") or rev.get("score")),
                "content": content,
                "date":    fmt_date(rev.get("date") or rev.get("at")),
                "reply":   clean_text(rev.get("reply_text") or rev.get("replyText")),
                "version": safe(rev.get("version")),
            })
        return out

    def _parse_signals(self, lst):
        out = []
        for s in (lst or []):
            content = clean_text(s.get("content"))
            if is_empty(content):
                continue
            out.append({
                "id":         safe(s.get("signal_id")),
                "type":       safe(s.get("signal_type")),
                "confidence": s.get("confidence"),
                "content":    content,
            })
        return out


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════════════

class ExcelBuilder:
    """Builds a polished, professional .xlsx workbook from ReportData."""

    # ── style constants ──────────────────────────────────────────────────────
    THIN   = Side(style="thin",   color=C_GREY_LINE)
    THICK  = Side(style="medium", color=C_ACCENT)
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _F(self, bold=False, size=10, color=C_TEXT_DARK, italic=False):
        return Font(name=FONT_NAME, bold=bold, size=size,
                    color=color, italic=italic)

    def _P(self, color):
        return PatternFill("solid", fgColor=color)

    def _A(self, h="left", v="top", wrap=True):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ── cell helpers ─────────────────────────────────────────────────────────

    def _write(self, ws, row, col, value, *,
               bold=False, size=10, fg=C_TEXT_DARK, italic=False,
               fill=None, halign="left", wrap=True, border=True,
               link=None, num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = self._F(bold, size, fg, italic)
        c.alignment = self._A(halign, "top", wrap)
        if fill:
            c.fill = self._P(fill)
        if border:
            c.border = self.BORDER
        if link:
            c.hyperlink = link
            c.font = Font(name=FONT_NAME, size=size, color=C_ACCENT,
                          underline="single", bold=bold)
        if num_fmt:
            c.number_format = num_fmt
        return c

    def _header_row(self, ws, row, labels, widths=None, col_start=1):
        for i, lbl in enumerate(labels):
            c = ws.cell(row=row, column=col_start + i, value=lbl)
            c.font      = self._F(bold=True, size=10, color=C_WHITE)
            c.fill      = self._P(C_DARK)
            c.alignment = self._A("center", "center", False)
            c.border    = self.BORDER
        if widths:
            for i, w in enumerate(widths):
                ws.column_dimensions[get_column_letter(col_start + i)].width = w

    def _section_banner(self, ws, row, title, span, col_start=1):
        c = ws.cell(row=row, column=col_start, value=f"  {title}")
        c.font      = self._F(bold=True, size=11, color=C_WHITE)
        c.fill      = self._P(C_ACCENT)
        c.alignment = self._A("left", "center", False)
        c.border    = self.BORDER
        if span > 1:
            ws.merge_cells(start_row=row, start_column=col_start,
                           end_row=row, end_column=col_start + span - 1)
        ws.row_dimensions[row].height = 22

    def _kv_row(self, ws, row, key, value, col=1, link=None):
        shade = C_GREY_BG if row % 2 == 0 else C_WHITE
        self._write(ws, row, col,   key,   bold=True, fill=shade)
        if link and is_url(str(value)):
            self._write(ws, row, col+1, str(value), fill=shade, link=str(value), wrap=True)
        else:
            self._write(ws, row, col+1, str(value), fill=shade, wrap=True)
        ws.row_dimensions[row].height = max(15, min(60, len(str(value)) // 3 + 15))

    def _sev_fill(self, sev_str):
        return SEV_PALETTE.get(sev_str.lower(), (C_TEXT_DARK, C_WHITE))[1]

    def _data_row(self, ws, row, values, shading=None):
        shade = shading or (C_GREY_BG if row % 2 == 0 else C_WHITE)
        for col, val in enumerate(values, 1):
            link = None
            if isinstance(val, tuple) and len(val) == 2:
                val, link = val
            self._write(ws, row, col, str(val) if val is not None else "—",
                        fill=shade, wrap=True, link=link)

    # ── sheet builders ───────────────────────────────────────────────────────

    def _sheet_overview(self, wb, d, logo_bytes):
        ws = wb.create_sheet("Overview")
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 52
        ws.freeze_panes = "A3"

        # Logo + company name header block
        ws.row_dimensions[1].height = 70
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value    = f"  {d.company_name}  —  Intelligence Report"
        title_cell.font     = self._F(bold=True, size=16, color=C_WHITE)
        title_cell.fill     = self._P(C_DARK)
        title_cell.alignment= self._A("left", "center", False)

        if logo_bytes:
            # tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            # tmp.write(logo_bytes); tmp.close()
            # img = XLImage(tmp.name)
            import tempfile
            import os
            tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            tmp.write(logo_bytes)
            tmp.flush()
            tmp.close()
            try:
                img = Image(tmp.name)
                # ... anchor and add image to sheet ...
                ws.add_image(img, 'A1')
                # ... rest of sheet setup ...
                wb.save(out_path)  # ← file still exists here ✓
            finally:
                os.unlink(tmp.name)

            img.width = 56; img.height = 56
            img.anchor = "A1"
            ws.add_image(img)
            os.unlink(tmp.name)

        ws.merge_cells("A2:B2")
        sub = ws["A2"]
        sub.value     = f"  {clean_domain(d.company_domain)}   ·   Generated {d.ingestion_date}"
        sub.font      = self._F(size=9, color="AAAAAA")
        sub.fill      = self._P(C_DARK)
        sub.alignment = self._A("left", "center", False)

        r = 3
        self._section_banner(ws, r, "Company Basics", 2); r += 1
        kv_pairs = [
            ("Company Name",      d.company_name,     False),
            ("Domain",            d.company_domain,   True),
            ("Year Founded",      d.year_founded,     False),
            ("HQ Location",       d.hq_location,      False),
            ("Operating In",      d.locations,        False),
            ("Industry",          d.industry,         False),
            ("Platforms",         d.platforms,        False),
            ("Employees",         d.employee_count,   False),
        ]
        for key, val, as_link in kv_pairs:
            if not is_empty(val):
                self._kv_row(ws, r, key, val, link=val if as_link else None); r += 1

        r += 1
        self._section_banner(ws, r, "Financials & Traction", 2); r += 1
        for key, val in [("Funding Stage", d.funding_stage), ("Funding Raised", d.funding_raised),
                         ("Annual Revenue", d.annual_revenue), ("No. of Users", d.no_of_users)]:
            if not is_empty(val):
                self._kv_row(ws, r, key, val); r += 1

        r += 1
        self._section_banner(ws, r, "Positioning", 2); r += 1
        for key, val in [("Key Positioning", d.key_positioning),
                         ("Revenue Model",   d.revenue_model)]:
            if not is_empty(val):
                self._kv_row(ws, r, key, val); r += 1

        if d.pricing_tiers:
            self._kv_row(ws, r, "Pricing Tiers", "\n".join(d.pricing_tiers)); r += 1
        if d.target_segments:
            self._kv_row(ws, r, "Target Segments", "  •  ".join(d.target_segments)); r += 1
        if d.tech_stack:
            self._kv_row(ws, r, "Tech Stack", ",  ".join(d.tech_stack)); r += 1

        r += 1
        self._section_banner(ws, r, "Official Links", 2); r += 1
        links = [("Play Store", d.playstore_link), ("App Store", d.appstore_link),
                 ("YouTube",    d.youtube_channel), ("LinkedIn",  d.linkedin_page)]
        for lbl, url in links:
            if not is_empty(url):
                self._kv_row(ws, r, lbl, url, link=url); r += 1

        r += 1
        self._section_banner(ws, r, "Market Sentiment", 2); r += 1
        ms = d.market_sentiment
        for key in ("overall", "analyst_view", "user_community_view"):
            val = ms.get(key)
            if not is_empty(val):
                self._kv_row(ws, r, key.replace("_", " ").title(), val); r += 1

        if d.csuite:
            r += 1
            self._section_banner(ws, r, "C-Suite Officers", 2); r += 1
            for officer in d.csuite:
                self._kv_row(ws, r, "", officer); r += 1

        if d.founders:
            r += 1
            self._section_banner(ws, r, "Founders", 2); r += 1
            self._kv_row(ws, r, "Founders", ",  ".join(d.founders)); r += 1

    def _sheet_app_store(self, wb, d):
        ws = wb.create_sheet("App Store")
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 30
        ws.freeze_panes = "B2"

        ws.row_dimensions[1].height = 22
        self._section_banner(ws, 1, "App Store Comparison", 3)
        self._header_row(ws, 2, ["Metric", "🤖 Google Play Store", " Apple App Store"])

        rows = [
            ("App Title",      d.ps.get("title"),      d.app.get("title")),
            ("Rating",         star_bar(d.ps.get("avg_rating") or d.ps.get("score")),
                               star_bar(d.app.get("score"))),
            ("Total Ratings",  d.ps.get("ratings"),    d.app.get("ratings")),
            ("Total Reviews",  d.ps.get("reviews"),    "—"),
            ("Installs",       d.ps.get("installs"),   "—"),
            ("Genre",          d.ps.get("genre"),      d.app.get("genre")),
            ("Price",          d.ps.get("free"),       d.app.get("price")),
            ("Version",        d.ps.get("version"),    d.app.get("version")),
            ("Released",       d.ps.get("released"),   d.app.get("released")),
            ("Content Rating", d.ps.get("content_rating"), d.app.get("content_rating")),
            ("Min OS",         d.ps.get("min_android"), d.app.get("min_ios")),
            ("Store URL",      d.ps.get("url"),         d.app.get("url")),
            ("Privacy Policy", d.ps.get("privacy_policy"), "—"),
        ]
        r = 3
        for lbl, pv, av in rows:
            if is_empty(pv) and is_empty(av):
                continue
            shade = C_GREY_BG if r % 2 == 0 else C_WHITE
            self._write(ws, r, 1, lbl, bold=True, fill=shade)
            pv_s = safe(pv)
            if is_url(pv_s):
                self._write(ws, r, 2, pv_s, fill=shade, link=pv_s, wrap=True)
            else:
                self._write(ws, r, 2, pv_s, fill=shade, wrap=True)
            av_s = safe(av)
            if is_url(av_s):
                self._write(ws, r, 3, av_s, fill=shade, link=av_s, wrap=True)
            else:
                self._write(ws, r, 3, av_s, fill=shade, wrap=True)
            r += 1

        # Rating distribution
        rd = d.ps.get("rating_dist", {})
        if rd:
            r += 1
            self._section_banner(ws, r, "Play Store Rating Distribution", 3); r += 1
            self._header_row(ws, r, ["Stars", "Count", "Visual"]); r += 1
            total = sum(int(v) for v in rd.values() if str(v).isdigit()) or 1
            for star in ["5", "4", "3", "2", "1"]:
                cnt = int(rd.get(star, rd.get(int(star), 0)) or 0)
                bar = "█" * int((cnt / total) * 20)
                shade = C_GREY_BG if r % 2 == 0 else C_WHITE
                self._write(ws, r, 1, f"{'⭐' * int(star)} {star} stars", fill=shade)
                self._write(ws, r, 2, cnt, fill=shade, halign="center")
                self._write(ws, r, 3, bar, fill=shade,
                            fg=C_AMBER if star in ("1","2") else C_GREEN)
                r += 1

        ws.freeze_panes = "A3"

    def _sheet_reviews(self, wb, d):
        reviews = d.ps.get("reviews_list", [])
        if not reviews:
            return
        ws = wb.create_sheet("Reviews")
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 55
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 12
        ws.freeze_panes = "A2"

        self._header_row(ws, 1, ["#", "Author", "★", "Review", "Developer Reply", "Date"])
        for i, rev in enumerate(reviews, 1):
            r = i + 1
            shade = C_GREY_BG if i % 2 == 0 else C_WHITE
            sev_shade = {1: C_RED_LIGHT, 2: C_AMBER_LIGHT}.get(
                int(safe(rev["rating"]) or 3), shade)
            self._write(ws, r, 1, i,                fill=sev_shade, halign="center")
            self._write(ws, r, 2, rev["author"],    fill=sev_shade)
            self._write(ws, r, 3, rev["rating"],    fill=sev_shade, halign="center")
            self._write(ws, r, 4, rev["content"],   fill=sev_shade, wrap=True)
            reply = rev["reply"] if not is_empty(rev["reply"]) else "—"
            self._write(ws, r, 5, reply,            fill=sev_shade, wrap=True,
                        italic=True, fg=C_TEXT_MUTED)
            self._write(ws, r, 6, rev["date"],      fill=sev_shade)
            ws.row_dimensions[r].height = max(15, min(80, len(rev["content"]) // 4 + 15))

    def _sheet_social(self, wb, d):
        # ── Reddit ────────────────────────────────────────────────────────────
        if d.reddit_posts:
            ws = wb.create_sheet("Reddit")
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 48
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 10
            ws.column_dimensions["F"].width = 18
            ws.freeze_panes = "A2"
            self._header_row(ws, 1, ["#", "Post Title", "Subreddit / Author", "Score", "Cmts", "Link"])
            for i, p in enumerate(d.reddit_posts, 1):
                r = i + 1
                shade = C_GREY_BG if i % 2 == 0 else C_WHITE
                self._write(ws, r, 1, i,                                      fill=shade, halign="center")
                self._write(ws, r, 2, p["title"],                             fill=shade, wrap=True)
                self._write(ws, r, 3, f"r/{p['subreddit']}  @{p['author']}", fill=shade)
                self._write(ws, r, 4, p["score"],                             fill=shade, halign="center")
                self._write(ws, r, 5, p["comments"],                          fill=shade, halign="center")
                url = p["url"]
                if is_url(url):
                    self._write(ws, r, 6, "Open ↗", fill=shade, link=url, halign="center")
                else:
                    self._write(ws, r, 6, "—", fill=shade, halign="center")

        # ── YouTube ───────────────────────────────────────────────────────────
        if d.youtube_videos:
            ws = wb.create_sheet("YouTube")
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 48
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 14
            ws.freeze_panes = "A2"
            self._header_row(ws, 1, ["#", "Video Title", "Views", "Likes", "Link"])
            for i, v in enumerate(d.youtube_videos, 1):
                r = i + 1
                shade = C_GREY_BG if i % 2 == 0 else C_WHITE
                self._write(ws, r, 1, i,          fill=shade, halign="center")
                self._write(ws, r, 2, v["title"], fill=shade, wrap=True)
                self._write(ws, r, 3, v["views"], fill=shade, halign="center")
                self._write(ws, r, 4, v["likes"], fill=shade, halign="center")
                url = v["url"]
                if is_url(url):
                    self._write(ws, r, 5, "Watch ↗", fill=shade, link=url, halign="center")
                else:
                    self._write(ws, r, 5, "—", fill=shade, halign="center")

    def _sheet_transcripts(self, wb, d):
        sigs = d.transcript.get("signals", [])
        if not sigs:
            return
        ws = wb.create_sheet("Transcript Signals")
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 62
        ws.freeze_panes = "A3"

        self._section_banner(ws, 1,
            f"Internal Transcript · {d.transcript.get('source_file')} · "
            f"{d.transcript.get('total_signals')} signals · "
            f"{d.transcript.get('meeting_type')}", 4)
        self._header_row(ws, 2, ["Signal ID", "Type", "Confidence", "Content"])

        type_colours = {
            "Trend":       C_ACCENT_LIGHT,
            "Risk":        C_RED_LIGHT,
            "Opportunity": C_GREEN_LIGHT,
            "Feature":     "E8F5E9",
            "Pain Point":  C_AMBER_LIGHT,
        }
        for i, s in enumerate(sigs, 1):
            r = i + 2
            shade = type_colours.get(s["type"], C_GREY_BG if i % 2 == 0 else C_WHITE)
            conf = s.get("confidence")
            conf_str = f"{float(conf):.0%}" if conf is not None else "—"
            self._write(ws, r, 1, s["id"],      fill=shade, halign="center")
            self._write(ws, r, 2, s["type"],    fill=shade, bold=True)
            self._write(ws, r, 3, conf_str,     fill=shade, halign="center")
            self._write(ws, r, 4, s["content"], fill=shade, wrap=True)
            ws.row_dimensions[r].height = max(15, min(90, len(s["content"]) // 5 + 15))

    def _sheet_problems(self, wb, d):
        if not d.problems:
            return
        ws = wb.create_sheet("Problems")
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 46
        ws.column_dimensions["C"].width = 11
        ws.column_dimensions["D"].width = 11
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 38
        ws.freeze_panes = "A3"

        self._section_banner(ws, 1,
            f"Identified Problems  —  Total: {d.total_problems}  |  "
            f"Top Categories: {d.top_categories}  |  "
            f"High Severity: {d.high_severity_count}", 7)
        self._header_row(ws, 2, ["ID", "Problem Statement", "Severity",
                                  "Frequency", "Category", "User Type", "Evidence Quotes"])
        for i, p in enumerate(d.problems, 1):
            r = i + 2
            shade = self._sev_fill(p["severity"])
            ev_str = "\n".join(f"• {e}" for e in p["evidence"]) if p["evidence"] else "—"
            self._write(ws, r, 1, p["id"],       fill=shade, halign="center", bold=True)
            self._write(ws, r, 2, p["problem"],  fill=shade, wrap=True)
            sev_txt = p["severity"]
            sev_fg  = SEV_PALETTE.get(sev_txt.lower(), (C_TEXT_DARK, C_WHITE))[0]
            c = ws.cell(row=r, column=3, value=sev_txt)
            c.font      = Font(name=FONT_NAME, size=10, bold=True, color=sev_fg)
            c.fill      = self._P(shade)
            c.alignment = self._A("center", "top", False)
            c.border    = self.BORDER
            self._write(ws, r, 4, p["frequency"],     fill=shade, halign="center")
            self._write(ws, r, 5, p["category"],      fill=shade)
            self._write(ws, r, 6, p["user_type"],     fill=shade)
            self._write(ws, r, 7, ev_str,             fill=shade, wrap=True, italic=True)
            ws.row_dimensions[r].height = max(20, min(100, len(p["problem"]) // 3 + 20))

    def _sheet_insights(self, wb, d):
        if not d.insights:
            return
        ws = wb.create_sheet("Strategic Insights")
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 44
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 36
        ws.column_dimensions["F"].width = 36
        ws.freeze_panes = "A3"

        self._section_banner(ws, 1,
            f"Strategic Insights  —  Total: {d.total_insights}  |  "
            f"Dominant Theme: {d.dominant_theme}  |  "
            f"Critical: {d.critical_count}", 6)
        self._header_row(ws, 2, ["ID", "Insight", "Priority", "Confidence",
                                  "Root Cause", "Implication"])
        for i, ins in enumerate(d.insights, 1):
            r = i + 2
            shade = self._sev_fill(ins["priority"])
            self._write(ws, r, 1, ins["id"],         fill=shade, halign="center", bold=True)
            self._write(ws, r, 2, ins["insight"],    fill=shade, wrap=True)
            pri_fg = SEV_PALETTE.get(ins["priority"].lower(), (C_TEXT_DARK, C_WHITE))[0]
            c = ws.cell(row=r, column=3, value=ins["priority"])
            c.font      = Font(name=FONT_NAME, size=10, bold=True, color=pri_fg)
            c.fill      = self._P(shade)
            c.alignment = self._A("center", "top", False)
            c.border    = self.BORDER
            self._write(ws, r, 4, ins["confidence"], fill=shade, halign="center")
            self._write(ws, r, 5, ins["root_cause"], fill=shade, wrap=True)
            self._write(ws, r, 6, ins["implication"],fill=shade, wrap=True)
            ws.row_dimensions[r].height = max(20, min(100, len(ins["insight"]) // 3 + 20))

        # Strategic summary block
        r = len(d.insights) + 4
        self._section_banner(ws, r, "Strategic Summary", 6); r += 1
        for lbl, val in [("Key Strategic Risk",   d.strategic_risk),
                         ("Biggest Opportunity",  d.biggest_opp)]:
            if not is_empty(val):
                shade = C_GREY_BG if r % 2 == 0 else C_WHITE
                self._write(ws, r, 1, lbl, bold=True, fill=shade)
                ws.merge_cells(start_row=r, start_column=2,
                               end_row=r, end_column=6)
                self._write(ws, r, 2, val, fill=shade, wrap=True); r += 1

    def _sheet_briefs(self, wb, d):
        if not d.briefs:
            return
        ws = wb.create_sheet("Product Briefs")
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 9
        ws.column_dimensions["D"].width = 9
        ws.column_dimensions["E"].width = 36
        ws.column_dimensions["F"].width = 36
        ws.column_dimensions["G"].width = 30
        ws.freeze_panes = "A3"

        self._section_banner(ws, 1,
            f"Product Briefs  —  Total: {d.total_briefs}  |  "
            f"Sprint Focus: {d.sprint_focus[:80]+'…' if len(d.sprint_focus) > 80 else d.sprint_focus}", 7)
        self._header_row(ws, 2, ["ID", "Feature Name", "Priority", "Effort",
                                  "Problem → Solution", "Expected Impact", "User Flow"])
        for i, b in enumerate(d.briefs, 1):
            r = i + 2
            shade = self._sev_fill(b["priority"])
            prob_sol = f"PROBLEM:\n{b['problem']}\n\nSOLUTION:\n{b['solution']}"
            flow = "\n".join(f"{j+1}. {s}" for j, s in enumerate(b["user_flow"])) if b["user_flow"] else "—"
            self._write(ws, r, 1, b["id"],      fill=shade, halign="center", bold=True)
            self._write(ws, r, 2, b["feature"], fill=shade, wrap=True, bold=True)
            pri_fg = SEV_PALETTE.get(b["priority"].lower(), (C_TEXT_DARK, C_WHITE))[0]
            c = ws.cell(row=r, column=3, value=b["priority"])
            c.font      = Font(name=FONT_NAME, size=10, bold=True, color=pri_fg)
            c.fill      = self._P(shade)
            c.alignment = self._A("center", "top", False)
            c.border    = self.BORDER
            self._write(ws, r, 4, b["effort"],  fill=shade, halign="center")
            self._write(ws, r, 5, prob_sol,     fill=shade, wrap=True)
            self._write(ws, r, 6, b["impact"],  fill=shade, wrap=True)
            self._write(ws, r, 7, flow,         fill=shade, wrap=True)
            ws.row_dimensions[r].height = max(30, min(120, len(prob_sol) // 4 + 20))

    def _sheet_company_detail(self, wb, d):
        ws = wb.create_sheet("Company Detail")
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60
        ws.freeze_panes = "A2"

        self._header_row(ws, 1, ["Attribute", "Detail"])

        def _block(title, rows):
            nonlocal r
            r += 1
            self._section_banner(ws, r, title, 2); r += 1
            for key, val in rows:
                if not is_empty(val):
                    self._kv_row(ws, r, key, val); r += 1

        r = 1
        _block("Milestones", [(f"Milestone {i+1}", m) for i, m in enumerate(d.milestones)])
        _block("New Features Launched", [(f"Feature {i+1}", f) for i, f in enumerate(d.new_features)])
        _block("Differentiators",
               [(item.get("feature", f"Item {i+1}"),
                 item.get("feature") or item.get("value", ""))
                for i, item in enumerate(d.differentiators)])
        _block("Competitors",
               [(item.get("name", f"#{i+1}"),
                 f"Domain: {item.get('domain','—')}")
                for i, item in enumerate(d.competitors)])
        _block("Strategic Moves",
               [(f"Move {i+1}", item.get("move") or item.get("value", ""))
                for i, item in enumerate(d.strategic_moves)])
        _block("Partnerships",
               [(item.get("partner", f"#{i+1}"),
                 f"{item.get('type','—')} · {item.get('description','—')} · {item.get('date','—')}")
                for i, item in enumerate(d.partnerships)])
        _block("User Complaints",
               [(f"Complaint {i+1}", item.get("issue") or item.get("value", ""))
                for i, item in enumerate(d.user_complaints)])
        _block("Regulatory & Legal",
               [(item.get("issue", f"Issue {i+1}"),
                 f"{item.get('status','—')} · {item.get('jurisdiction','—')} · {item.get('date','—')}")
                for i, item in enumerate(d.regulatory)])
        _block("Current Problems / Struggles",
               [(f"Problem {i+1}", item.get("description") or item.get("value",""))
                for i, item in enumerate(d.current_problems)])
        _block("Other Crucial Details", [(f"Detail {i+1}", t) for i, t in enumerate(d.other_details)])

    def build(self, d, logo_bytes, out_path):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._sheet_overview(wb, d, logo_bytes)
        self._sheet_app_store(wb, d)
        self._sheet_reviews(wb, d)
        self._sheet_social(wb, d)
        self._sheet_transcripts(wb, d)
        self._sheet_problems(wb, d)
        self._sheet_insights(wb, d)
        self._sheet_briefs(wb, d)
        self._sheet_company_detail(wb, d)

        wb.save(out_path)
        print(f"  ✓ Excel  →  {out_path}")


# ══════════════════════════════════════════════════════════════════════════════
# DOCX BUILDER  (pure Python via subprocess calling node generate_docx.js)
# We emit a Node.js script inline to keep everything in one Python file.
# ══════════════════════════════════════════════════════════════════════════════

DOCX_JS_TEMPLATE = r'''
"use strict";
const fs   = require("fs");
const path = require("path");
const {
  Document, Packer, Paragraph, TextRun, Table, TableRow, TableCell,
  ExternalHyperlink, HeadingLevel, AlignmentType, LevelFormat,
  BorderStyle, WidthType, ShadingType, VerticalAlign,
  PageNumber, PageBreak, Header, Footer, ImageRun,
  TabStopType, TabStopPosition,
} = require("docx");

// ── data injected by Python ───────────────────────────────────────────
const D          = JSON.parse(fs.readFileSync("__DATA_JSON__","utf8"));
const LOGO_FILE  = "__LOGO_FILE__";
const OUT_PATH   = "__OUT_PATH__";

// ── palette ───────────────────────────────────────────────────────────
const FONT  = "Calibri";
const CDARK = "0D1F3C";
const CACC  = "1A6BCC";
const CACCL = "D6E8FF";
const CGREY = "F5F7FA";
const CLINE = "DEE2E8";
const CWHITE= "FFFFFF";
const CMUTED= "6B7280";
const SEV_COLOURS = {
  critical: { text:"B71C1C", bg:"FDECEA" },
  high:     { text:"B76E00", bg:"FFF3CD" },
  medium:   { text:"7B5800", bg:"FFFBEA" },
  low:      { text:"1A7A4A", bg:"D4EDDA" },
  p0:       { text:"B71C1C", bg:"FDECEA" },
  p1:       { text:"B76E00", bg:"FFF3CD" },
  p2:       { text:"7B5800", bg:"FFFBEA" },
  p3:       { text:"1A7A4A", bg:"D4EDDA" },
};
function sevColour(s) {
  return SEV_COLOURS[(s||"").toLowerCase()] || { text: CDARK, bg: CWHITE };
}

// ── helpers ───────────────────────────────────────────────────────────
const safe = (v,fb="—") => {
  if(v===null||v===undefined||v==="")return fb;
  if(Array.isArray(v)){ const f=v.filter(x=>x!==null&&x!==""); return f.length?f.join(" • "):fb; }
  return String(v);
};
const isUrl = v => typeof v==="string" && /^https?:\/\//.test(v);
const noEmpty = v => v!==null&&v!==undefined&&v!==""&&v!=="—";

const CELL_BORDER = {
  top:    {style:BorderStyle.SINGLE,size:4,color:CLINE},
  bottom: {style:BorderStyle.SINGLE,size:4,color:CLINE},
  left:   {style:BorderStyle.SINGLE,size:4,color:CLINE},
  right:  {style:BorderStyle.SINGLE,size:4,color:CLINE},
};
const CELL_MARGINS = {top:80,bottom:80,left:120,right:120};

// text runs
const run = (text,opts={}) => new TextRun({text:safe(text),font:FONT,size:20,...opts});
const bold = (text,opts={}) => run(text,{bold:true,...opts});
const muted= (text)          => run(text,{color:CMUTED,italics:true});
const link = (text,url)      => new ExternalHyperlink({link:url,
  children:[new TextRun({text,font:FONT,size:20,color:CACC,underline:{}})]});

// paragraphs
const para = (children,opts={}) =>
  new Paragraph({children:Array.isArray(children)?children:[children],...opts});
const blank = () => para([run("")]);
const h1 = t => new Paragraph({heading:HeadingLevel.HEADING_1,children:[bold(t,{size:28,color:CDARK})]});
const h2 = t => new Paragraph({heading:HeadingLevel.HEADING_2,children:[bold(t,{size:24,color:CACC})]});
const h3 = t => new Paragraph({heading:HeadingLevel.HEADING_3,children:[bold(t,{size:21,color:"2C3E50"})]});
const pageBreak = () => new Paragraph({children:[new PageBreak()]});
const bullet = (text,opts={}) => new Paragraph({
  numbering:{reference:"bullets",level:0},
  children:[run(text),...(opts.extra||[])],
});

// tables
const hdrCell = (text,w) => new TableCell({
  width:{size:w,type:WidthType.DXA}, shading:{fill:CDARK,type:ShadingType.CLEAR},
  borders:CELL_BORDER, margins:CELL_MARGINS,
  children:[para([bold(text,{color:CWHITE})])],
});
const cell = (children,w,shade=CWHITE) => new TableCell({
  width:{size:w,type:WidthType.DXA}, shading:{fill:shade,type:ShadingType.CLEAR},
  borders:CELL_BORDER, margins:CELL_MARGINS,
  children:Array.isArray(children)?children:[para([typeof children==="string"?run(children):children])],
});
const strCell  = (t,w,s)   => cell([para([run(safe(t))])],w,s);
const boldCell = (t,w,s)   => cell([para([bold(safe(t))])],w,s);
const linkCell = (t,u,w,s) => isUrl(u)
  ? cell([para([link(safe(t),u)])],w,s)
  : strCell(t,w,s);

// shaded KV table — full width 9360
function kvTable(pairs) {
  const rows = pairs
    .filter(([,v])=>noEmpty(v)&&v!=="—")
    .map(([k,v,asLink],i)=>{
      const shade = i%2===0?CGREY:CWHITE;
      return new TableRow({children:[
        new TableCell({width:{size:2640,type:WidthType.DXA},
          shading:{fill:i%2===0?"E8F0FC":CACCL,type:ShadingType.CLEAR},
          borders:CELL_BORDER, margins:CELL_MARGINS,
          children:[para([bold(k,{size:19})])]}),
        new TableCell({width:{size:6720,type:WidthType.DXA},
          shading:{fill:shade,type:ShadingType.CLEAR},
          borders:CELL_BORDER, margins:CELL_MARGINS,
          children: asLink&&isUrl(safe(v))
            ? [para([link(safe(v),safe(v))])]
            : [para([run(safe(v))])]}),
      ]});
    });
  if(!rows.length) return null;
  return new Table({width:{size:9360,type:WidthType.DXA},columnWidths:[2640,6720],rows});
}

// accent divider paragraph
const divider = (label) => new Paragraph({
  border:{bottom:{style:BorderStyle.SINGLE,size:8,color:CACC,space:1}},
  spacing:{before:240,after:120},
  children:[bold(label,{size:22,color:CACC})],
});

// ── numbering & styles ────────────────────────────────────────────────
const numbering = {config:[
  {reference:"bullets",levels:[{level:0,format:LevelFormat.BULLET,text:"\u2022",
    alignment:AlignmentType.LEFT,
    style:{paragraph:{indent:{left:720,hanging:360}}}}]},
  {reference:"numbers",levels:[{level:0,format:LevelFormat.DECIMAL,text:"%1.",
    alignment:AlignmentType.LEFT,
    style:{paragraph:{indent:{left:720,hanging:360}}}}]},
]};

const styles = {
  default:{document:{run:{font:FONT,size:20}}},
  paragraphStyles:[
    {id:"Heading1",name:"Heading 1",basedOn:"Normal",next:"Normal",quickFormat:true,
     run:{size:32,bold:true,font:FONT,color:CDARK},
     paragraph:{spacing:{before:360,after:180},outlineLevel:0}},
    {id:"Heading2",name:"Heading 2",basedOn:"Normal",next:"Normal",quickFormat:true,
     run:{size:26,bold:true,font:FONT,color:CACC},
     paragraph:{spacing:{before:240,after:120},outlineLevel:1}},
    {id:"Heading3",name:"Heading 3",basedOn:"Normal",next:"Normal",quickFormat:true,
     run:{size:22,bold:true,font:FONT,color:"2C3E50"},
     paragraph:{spacing:{before:180,after:90},outlineLevel:2}},
  ]
};

// ══════════════════════════════════════════════════════════════════════
// BUILD CONTENT
// ══════════════════════════════════════════════════════════════════════

const children = [];

// ── COVER ────────────────────────────────────────────────────────────
const cp       = D.company_profile || {};
const compName = safe(cp.company_name || D.project_name);
const compDom  = safe(cp.domain || D.domain);

// Logo image (if available)
let logoImage = null;
if(LOGO_FILE && fs.existsSync(LOGO_FILE)){
  try{
    logoImage = new Paragraph({
      alignment:AlignmentType.LEFT,
      spacing:{before:0,after:240},
      children:[new ImageRun({
        type:"png",
        data: fs.readFileSync(LOGO_FILE),
        transformation:{width:72,height:72},
        altText:{title:"Logo",description:"Company Logo",name:"Logo"},
      })],
    });
  }catch(e){}
}

// Cover block
if(logoImage) children.push(logoImage);
children.push(
  new Paragraph({
    spacing:{before:logoImage?120:480,after:120},
    children:[bold(compName,{size:52,color:CDARK})],
  }),
  new Paragraph({
    spacing:{before:0,after:80},
    children:[run(compDom,{size:24,color:CACC})],
  }),
  new Paragraph({
    border:{bottom:{style:BorderStyle.SINGLE,size:12,color:CACC,space:1}},
    spacing:{before:0,after:360},
    children:[run("")],
  }),
);

// Two-column cover summary table (no header)
const coverKV = [
  ["Industry",      cp.industry_and_segment],
  ["Founded",       cp.year_founded],
  ["Headquarters",  cp.exact_hq_location],
  ["Employees",     cp.employee_count],
  ["Revenue",       cp.annual_revenue],
  ["Funding",       `${safe(cp.funding_stage)} · ${safe(cp.funding_raised)}`],
  ["Platforms",     cp.available_platforms],
  ["Revenue Model", cp.revenue_model],
].filter(([,v])=>noEmpty(v)&&v!=="—"&&v!=="— · —");

if(coverKV.length){
  const leftRows  = coverKV.slice(0, Math.ceil(coverKV.length/2));
  const rightRows = coverKV.slice(Math.ceil(coverKV.length/2));
  const maxR = Math.max(leftRows.length, rightRows.length);
  const tblRows = [];
  for(let i=0;i<maxR;i++){
    const [lk,lv] = leftRows[i]  || ["",""];
    const [rk,rv] = rightRows[i] || ["",""];
    tblRows.push(new TableRow({children:[
      new TableCell({width:{size:2160,type:WidthType.DXA},
        shading:{fill:"E8F0FC",type:ShadingType.CLEAR},
        borders:CELL_BORDER,margins:CELL_MARGINS,
        children:[para([bold(safe(lk),{size:18})])]}),
      new TableCell({width:{size:2520,type:WidthType.DXA},
        shading:{fill:CWHITE,type:ShadingType.CLEAR},
        borders:CELL_BORDER,margins:CELL_MARGINS,
        children:[para([run(safe(lv),{size:18})])]}),
      new TableCell({width:{size:180,type:WidthType.DXA},
        shading:{fill:CWHITE,type:ShadingType.CLEAR},
        borders:{top:{style:BorderStyle.NONE},bottom:{style:BorderStyle.NONE},
                 left:{style:BorderStyle.NONE},right:{style:BorderStyle.NONE}},
        children:[para([run("")])]}),
      new TableCell({width:{size:2160,type:WidthType.DXA},
        shading:{fill:"E8F0FC",type:ShadingType.CLEAR},
        borders:CELL_BORDER,margins:CELL_MARGINS,
        children:[para([bold(safe(rk),{size:18})])]}),
      new TableCell({width:{size:2340,type:WidthType.DXA},
        shading:{fill:CWHITE,type:ShadingType.CLEAR},
        borders:CELL_BORDER,margins:CELL_MARGINS,
        children:[para([run(safe(rv),{size:18})])]}),
    ]}));
  }
  children.push(
    new Table({width:{size:9360,type:WidthType.DXA},
      columnWidths:[2160,2520,180,2160,2340],rows:tblRows}),
    blank()
  );
}

// Key Positioning callout
if(noEmpty(cp.key_positioning)){
  children.push(
    new Paragraph({
      shading:{fill:CACCL,type:ShadingType.CLEAR},
      border:{
        left:{style:BorderStyle.SINGLE,size:24,color:CACC,space:6},
      },
      spacing:{before:240,after:240},
      indent:{left:300,right:300},
      children:[
        bold("Key Positioning:  ",{color:CACC}),
        run(safe(cp.key_positioning),{italics:true}),
      ],
    }),
    blank()
  );
}
children.push(pageBreak());

// ── 1. COMPANY PROFILE ───────────────────────────────────────────────
children.push(h1("Company Profile"), blank());

const officialLinks = [
  ["Play Store",   cp.playstore_link],
  ["App Store",    cp.appstore_link],
  ["YouTube",      cp.youtube_official_channel],
  ["LinkedIn",     cp.linkedin_company_page],
].filter(([,v])=>isUrl(safe(v)));

if(officialLinks.length){
  children.push(h2("Official Links"), blank());
  children.push(new Table({
    width:{size:9360,type:WidthType.DXA},
    columnWidths:[1800,7560],
    rows: officialLinks.map(([lbl,url],i)=>new TableRow({children:[
      new TableCell({width:{size:1800,type:WidthType.DXA},
        shading:{fill:i%2===0?"E8F0FC":CACCL,type:ShadingType.CLEAR},
        borders:CELL_BORDER,margins:CELL_MARGINS,
        children:[para([bold(lbl,{size:19})])]}),
      new TableCell({width:{size:7560,type:WidthType.DXA},
        shading:{fill:i%2===0?CGREY:CWHITE,type:ShadingType.CLEAR},
        borders:CELL_BORDER,margins:CELL_MARGINS,
        children:[para([link(url,url)])]}),
    ]}))
  }));
  children.push(blank());
}

// Leadership
const csuite = (cp["c-suite_officer"]||cp.c_suite_officer||[]).filter(noEmpty);
const founders = (cp.names_of_founders||[]).filter(noEmpty);
if(csuite.length||founders.length){
  children.push(h2("Leadership"), blank());
  const leaderRows = [];
  csuite.forEach((o,i)=>leaderRows.push(new TableRow({children:[
    new TableCell({width:{size:1800,type:WidthType.DXA},
      shading:{fill:i%2===0?"E8F0FC":CACCL,type:ShadingType.CLEAR},
      borders:CELL_BORDER,margins:CELL_MARGINS,
      children:[para([bold("Executive",{size:18})])]}),
    new TableCell({width:{size:7560,type:WidthType.DXA},
      shading:{fill:i%2===0?CGREY:CWHITE,type:ShadingType.CLEAR},
      borders:CELL_BORDER,margins:CELL_MARGINS,
      children:[para([run(safe(o))])]}),
  ]})));
  if(founders.length) leaderRows.push(new TableRow({children:[
    new TableCell({width:{size:1800,type:WidthType.DXA},
      shading:{fill:"E8F0FC",type:ShadingType.CLEAR},
      borders:CELL_BORDER,margins:CELL_MARGINS,
      children:[para([bold("Founders",{size:18})])]}),
    new TableCell({width:{size:7560,type:WidthType.DXA},
      shading:{fill:CWHITE,type:ShadingType.CLEAR},
      borders:CELL_BORDER,margins:CELL_MARGINS,
      children:[para([run(founders.join("  •  "))])]}),
  ]}));
  if(leaderRows.length)
    children.push(new Table({width:{size:9360,type:WidthType.DXA},
      columnWidths:[1800,7560],rows:leaderRows}),blank());
}

// Pricing
const pricing = (cp.pricing_tiers||[]).filter(noEmpty);
if(pricing.length){
  children.push(h2("Pricing Tiers"), blank());
  pricing.forEach(t=>children.push(bullet(t)));
  children.push(blank());
}

// Target Segments
const segs = (cp.target_customer_segments||[]).filter(noEmpty);
if(segs.length){
  children.push(h2("Target Customer Segments"), blank());
  segs.forEach(s=>children.push(bullet(s)));
  children.push(blank());
}

// Tech Stack
const tech = (cp.tech_stack_highlights||[]).filter(noEmpty);
if(tech.length){
  children.push(h2("Technology Highlights"), blank());
  tech.forEach(t=>children.push(bullet(t)));
  children.push(blank());
}

// Market Sentiment
const ms = cp.market_sentiment||{};
if(ms.overall||ms.analyst_view){
  children.push(h2("Market Sentiment"), blank());
  const msT = kvTable([
    ["Overall",            ms.overall],
    ["Analyst View",       ms.analyst_view],
    ["Community View",     ms.user_community_view],
    ["As of",              ms.date],
  ]);
  if(msT){ children.push(msT,blank()); }
}
children.push(pageBreak());

// ── 2. APP STORE ANALYSIS ────────────────────────────────────────────
const ps  = D.play_store  || {};
const app = D.app_store   || {};
if(Object.keys(ps).length || Object.keys(app).length){
  children.push(h1("App Store Analysis"), blank());

  // Side-by-side comparison table
  const psMeta  = ps.score||ps.avg_rating   ? ps  : {};
  const appMeta = app.score                  ? app : {};
  const comparisons = [
    ["App Title",      psMeta.title,       appMeta.title],
    ["Rating",         psMeta.score ? `${parseFloat(psMeta.score||0).toFixed(1)} ★` : "—",
                       appMeta.score ? `${parseFloat(appMeta.score||0).toFixed(1)} ★` : "—"],
    ["Total Ratings",  psMeta.ratings,     appMeta.ratings],
    ["Installs",       psMeta.installs,    "—"],
    ["Genre",          psMeta.genre,       appMeta.genre],
    ["Price",          psMeta.free,        appMeta.price],
    ["Released",       psMeta.released,    appMeta.released],
    ["Version",        psMeta.version,     appMeta.version],
    ["Content Rating", psMeta.content_rating, appMeta.content_rating],
  ].filter(([,a,b])=>noEmpty(a)||noEmpty(b));

  if(comparisons.length){
    children.push(h2("Play Store vs App Store"), blank());
    children.push(new Table({
      width:{size:9360,type:WidthType.DXA},
      columnWidths:[2400,3480,3480],
      rows:[
        new TableRow({children:[
          hdrCell("Metric",2400),
          hdrCell("🤖 Google Play",3480),
          hdrCell(" Apple App Store",3480),
        ]}),
        ...comparisons.map(([lbl,pv,av],i)=>{
          const shade = i%2===0?CGREY:CWHITE;
          return new TableRow({children:[
            boldCell(lbl,2400,i%2===0?"E8F0FC":CACCL),
            strCell(safe(pv),3480,shade),
            strCell(safe(av),3480,shade),
          ]});
        })
      ]
    }));
    children.push(blank());
  }

  // Rating distribution
  const rd = ps.rating_dist||{};
  const rdEntries = Object.entries(rd).filter(([,v])=>noEmpty(v));
  if(rdEntries.length){
    children.push(h2("Play Store Rating Distribution"), blank());
    const total = rdEntries.reduce((s,[,v])=>s+parseInt(v||0),0)||1;
    children.push(new Table({
      width:{size:9360,type:WidthType.DXA},
      columnWidths:[1440,1440,6480],
      rows:[
        new TableRow({children:[hdrCell("Stars",1440),hdrCell("Count",1440),hdrCell("Distribution",6480)]}),
        ...["5","4","3","2","1"].map((star,i)=>{
          const cnt = parseInt(rd[star]||rd[parseInt(star)]||0);
          const bar = "█".repeat(Math.round((cnt/total)*30));
          const shade = i%2===0?CGREY:CWHITE;
          const barColour = ["1","2"].includes(star) ? "B71C1C" : "1A7A4A";
          return new TableRow({children:[
            strCell(`${"⭐".repeat(parseInt(star))} ${star}★`,1440,shade),
            strCell(String(cnt),1440,shade),
            cell([para([run(bar||"—",{color:barColour})])],6480,shade),
          ]});
        })
      ]
    }));
    children.push(blank());
  }

  // Review summary
  const reviews = D.reviews_list||[];
  if(reviews.length){
    children.push(h2(`Top Reviews (${reviews.length} analysed)`), blank());
    // Show first 5 most impactful (1 and 2 star first, then rest)
    const sorted = [...reviews].sort((a,b)=>parseInt(a.rating||5)-parseInt(b.rating||5));
    sorted.slice(0,8).forEach((rev,i)=>{
      const sc = parseInt(rev.rating||3);
      const sev = sc<=2?"critical":sc<=3?"medium":"low";
      const col = SEV_COLOURS[sev];
      children.push(
        new Paragraph({
          shading:{fill:col.bg,type:ShadingType.CLEAR},
          border:{left:{style:BorderStyle.SINGLE,size:20,color:col.text,space:4}},
          spacing:{before:120,after:0},
          indent:{left:240},
          children:[
            bold(`${"⭐".repeat(sc)}  ${safe(rev.author)}`,{color:col.text}),
            run(`  ·  ${safe(rev.date)}`,{color:CMUTED,size:18}),
          ],
        }),
        new Paragraph({
          shading:{fill:col.bg,type:ShadingType.CLEAR},
          spacing:{before:0,after:rev.reply&&noEmpty(rev.reply)?0:160},
          indent:{left:240,right:240},
          children:[run(safe(rev.content))],
        }),
      );
      if(rev.reply && noEmpty(rev.reply)){
        children.push(new Paragraph({
          shading:{fill:"F0F4FF",type:ShadingType.CLEAR},
          spacing:{before:0,after:160},
          indent:{left:360,right:240},
          children:[bold("Developer: ",{color:CACC,size:18}),
                    run(safe(rev.reply),{italics:true,size:18,color:CMUTED})],
        }));
      }
    });
    children.push(blank());
  }
  children.push(pageBreak());
}

// ── 3. SOCIAL INTELLIGENCE ───────────────────────────────────────────
const reddit  = D.reddit_posts   || [];
const youtube = D.youtube_videos || [];
if(reddit.length || youtube.length){
  children.push(h1("Social & Media Intelligence"), blank());

  if(reddit.length){
    children.push(h2(`Reddit Posts  (${reddit.length})`), blank());
    children.push(new Table({
      width:{size:9360,type:WidthType.DXA},
      columnWidths:[5640,1080,1080,1560],
      rows:[
        new TableRow({children:[hdrCell("Post Title",5640),hdrCell("Subreddit",1080),
                                hdrCell("Score",1080),hdrCell("Link",1560)]}),
        ...reddit.map((p,i)=>{
          const shade = i%2===0?CGREY:CWHITE;
          return new TableRow({children:[
            strCell(safe(p.title),5640,shade),
            strCell(`r/${safe(p.subreddit)}`,1080,shade),
            strCell(safe(p.score),1080,shade),
            isUrl(safe(p.url))
              ? cell([para([link("Open ↗",safe(p.url))])],1560,shade)
              : strCell("—",1560,shade),
          ]});
        })
      ]
    }));
    children.push(blank());
  }

  if(youtube.length){
    children.push(h2(`YouTube Coverage  (${youtube.length} videos)`), blank());
    children.push(new Table({
      width:{size:9360,type:WidthType.DXA},
      columnWidths:[6120,1080,1080,1080],
      rows:[
        new TableRow({children:[hdrCell("Video Title",6120),hdrCell("Views",1080),
                                hdrCell("Likes",1080),hdrCell("Link",1080)]}),
        ...youtube.map((v,i)=>{
          const shade = i%2===0?CGREY:CWHITE;
          return new TableRow({children:[
            strCell(safe(v.title),6120,shade),
            strCell(safe(v.views),1080,shade),
            strCell(safe(v.likes),1080,shade),
            isUrl(safe(v.url))
              ? cell([para([link("Watch ↗",safe(v.url))])],1080,shade)
              : strCell("—",1080,shade),
          ]});
        })
      ]
    }));
    children.push(blank());
  }
  children.push(pageBreak());
}

// ── 4. TRANSCRIPT SIGNALS ────────────────────────────────────────────
const signals = D.signals||[];
if(signals.length){
  children.push(h1("Internal Transcript Signals"), blank());
  const trMeta = D.transcript_meta||{};
  const trT = kvTable([
    ["Source File",   trMeta.source_file],
    ["Total Signals", trMeta.total_signals],
    ["Classifier",    trMeta.classifier],
    ["Meeting Type",  trMeta.meeting_type],
  ]);
  if(trT){children.push(trT,blank());}

  const TYPE_COLOUR = {
    "Trend":"1A6BCC","Risk":"B71C1C","Opportunity":"1A7A4A",
    "Feature":"2C7A4B","Pain Point":"B76E00",
  };
  signals.forEach((s,i)=>{
    const col = TYPE_COLOUR[s.type]||CDARK;
    const conf = s.confidence!=null ? `${(parseFloat(s.confidence)*100).toFixed(0)}%` : "—";
    children.push(new Paragraph({
      shading:{fill:"F4F6FA",type:ShadingType.CLEAR},
      border:{left:{style:BorderStyle.SINGLE,size:16,color:col,space:4}},
      spacing:{before:120,after:0},
      indent:{left:240},
      children:[
        bold(`[${safe(s.id)}]  ${safe(s.type)}`,{color:col}),
        run(`  ·  Confidence: ${conf}`,{color:CMUTED,size:18}),
      ],
    }),
    new Paragraph({
      shading:{fill:"F4F6FA",type:ShadingType.CLEAR},
      spacing:{before:0,after:160},
      indent:{left:240,right:240},
      children:[run(safe(s.content),{size:18})],
    }));
  });
  children.push(pageBreak());
}

// ── 5. IDENTIFIED PROBLEMS ───────────────────────────────────────────
const problems = D.problems||[];
if(problems.length){
  children.push(h1("Identified Problems"), blank());
  // Summary stats row
  children.push(new Table({
    width:{size:9360,type:WidthType.DXA},
    columnWidths:[2340,2340,2340,2340],
    rows:[new TableRow({children:[
      new TableCell({width:{size:2340,type:WidthType.DXA},
        shading:{fill:CDARK,type:ShadingType.CLEAR},borders:CELL_BORDER,margins:CELL_MARGINS,
        children:[para([bold("Total Problems",{color:CWHITE,size:18})]),
                  para([run(safe(D.total_problems||problems.length),{color:CWHITE,size:28,bold:true})])]}),
      new TableCell({width:{size:2340,type:WidthType.DXA},
        shading:{fill:"FDECEA",type:ShadingType.CLEAR},borders:CELL_BORDER,margins:CELL_MARGINS,
        children:[para([bold("Critical / High",{color:"B71C1C",size:18})]),
                  para([run(safe(D.high_severity_count||"—"),{color:"B71C1C",size:28,bold:true})])]}),
      new TableCell({width:{size:2340,type:WidthType.DXA},
        shading:{fill:CACCL,type:ShadingType.CLEAR},borders:CELL_BORDER,margins:CELL_MARGINS,
        children:[para([bold("Top Categories",{color:CACC,size:18})]),
                  para([run(safe(D.top_categories),{color:CACC,size:18})])]}),
      new TableCell({width:{size:2340,type:WidthType.DXA},
        shading:{fill:"D4EDDA",type:ShadingType.CLEAR},borders:CELL_BORDER,margins:CELL_MARGINS,
        children:[para([bold("Sources Used",{color:"1A7A4A",size:18})]),
                  para([run(safe(D.sources_used),{color:"1A7A4A",size:18})])]}),
    ]})]
  }));
  children.push(blank());

  problems.forEach((p,i)=>{
    const col = sevColour(p.severity);
    children.push(
      new Paragraph({
        spacing:{before:240,after:0},
        children:[
          bold(`${safe(p.id)} — `,{color:CDARK,size:22}),
          run(safe(p.severity).toUpperCase(),{color:col.text,bold:true,size:20}),
          run(`  ·  ${safe(p.category)}  ·  ${safe(p.frequency)}`,{color:CMUTED,size:18}),
        ],
      }),
      new Paragraph({
        shading:{fill:col.bg,type:ShadingType.CLEAR},
        border:{left:{style:BorderStyle.SINGLE,size:20,color:col.text,space:4}},
        spacing:{before:80,after:100},
        indent:{left:240,right:160},
        children:[run(safe(p.problem))],
      })
    );
    // Evidence
    const ev = p.evidence||[];
    if(ev.filter(noEmpty).length){
      children.push(para([bold("Evidence:",{color:CDARK,size:18})]));
      ev.filter(noEmpty).forEach(e=>children.push(bullet(`"${e}"`,{})));
    }
    children.push(blank());
  });
  children.push(pageBreak());
}

// ── 6. STRATEGIC INSIGHTS ────────────────────────────────────────────
const insights = D.insights||[];
if(insights.length){
  children.push(h1("Strategic Insights"), blank());

  // Strategic summary callouts
  if(noEmpty(D.strategic_risk)){
    children.push(new Paragraph({
      shading:{fill:"FDECEA",type:ShadingType.CLEAR},
      border:{left:{style:BorderStyle.SINGLE,size:24,color:"B71C1C",space:6}},
      spacing:{before:0,after:80},
      indent:{left:240,right:240},
      children:[bold("⚠  Strategic Risk: ",{color:"B71C1C"}),run(safe(D.strategic_risk))],
    }));
  }
  if(noEmpty(D.biggest_opp)){
    children.push(new Paragraph({
      shading:{fill:"D4EDDA",type:ShadingType.CLEAR},
      border:{left:{style:BorderStyle.SINGLE,size:24,color:"1A7A4A",space:6}},
      spacing:{before:0,after:200},
      indent:{left:240,right:240},
      children:[bold("✓  Opportunity: ",{color:"1A7A4A"}),run(safe(D.biggest_opp))],
    }));
  }

  insights.forEach((ins,i)=>{
    const col = sevColour(ins.priority);
    children.push(
      new Paragraph({
        spacing:{before:240,after:0},
        children:[
          bold(`${safe(ins.id)} — `,{color:CDARK,size:22}),
          run(safe(ins.theme).toUpperCase(),{color:col.text,bold:true,size:20}),
          run(`  ·  ${safe(ins.priority)}  ·  Confidence: ${safe(ins.confidence)}`,{color:CMUTED,size:18}),
        ],
      }),
      new Paragraph({
        shading:{fill:col.bg,type:ShadingType.CLEAR},
        border:{left:{style:BorderStyle.SINGLE,size:20,color:col.text,space:4}},
        spacing:{before:80,after:100},
        indent:{left:240,right:160},
        children:[run(safe(ins.insight))],
      })
    );
    const insDetail = kvTable([
      ["Root Cause",      ins.root_cause],
      ["Evidence",        ins.evidence],
      ["Competitor Gap",  ins.competitor_gap],
      ["Opportunity",     ins.opportunity],
      ["Implication",     ins.implication],
    ]);
    if(insDetail){children.push(insDetail);}
    children.push(blank());
  });
  children.push(pageBreak());
}

// ── 7. PRODUCT BRIEFS ────────────────────────────────────────────────
const briefs = D.briefs||[];
if(briefs.length){
  children.push(h1("Product Briefs"), blank());

  if(noEmpty(D.sprint_focus)){
    children.push(new Paragraph({
      shading:{fill:CACCL,type:ShadingType.CLEAR},
      border:{left:{style:BorderStyle.SINGLE,size:24,color:CACC,space:6}},
      spacing:{before:0,after:240},
      indent:{left:240,right:240},
      children:[bold("Recommended Sprint Focus: ",{color:CACC}),run(safe(D.sprint_focus))],
    }));
  }

  briefs.forEach((b,i)=>{
    const col = sevColour(b.priority);
    children.push(
      new Paragraph({
        spacing:{before:240,after:0},
        children:[
          bold(`${safe(b.id)}  `,{size:22,color:CDARK}),
          run(safe(b.feature),{size:22,bold:true,color:CACC}),
          run(`  [${safe(b.priority)}]  ·  Effort: ${safe(b.effort)}`,{color:CMUTED,size:18}),
        ],
      })
    );
    // Brief detail table
    const briefT = kvTable([
      ["Problem",   b.problem],
      ["Why Now",   b.why_now],
      ["Solution",  b.solution],
      ["Impact",    b.impact],
      ["Metric",    b.metric],
    ]);
    if(briefT){children.push(briefT);}

    // User flow
    const uf = (b.user_flow||[]).filter(noEmpty);
    if(uf.length){
      children.push(blank(),para([bold("User Flow:",{color:CDARK,size:18})]));
      uf.forEach((step,si)=>children.push(new Paragraph({
        numbering:{reference:"numbers",level:0},
        children:[run(safe(step))],
      })));
    }
    children.push(blank());
  });
  children.push(pageBreak());
}

// ── 8. MILESTONES & STRATEGIC MOVES ──────────────────────────────────
const milestones = D.milestones||[];
const moves      = D.strategic_moves||[];
const partners   = D.partnerships||[];
if(milestones.length||moves.length||partners.length){
  children.push(h1("Timeline & Strategy"), blank());
  if(milestones.length){
    children.push(h2("Key Milestones"), blank());
    milestones.forEach(m=>children.push(bullet(safe(typeof m==="object"?m.value||JSON.stringify(m):m))));
    children.push(blank());
  }
  if(moves.length){
    children.push(h2("Strategic Moves"), blank());
    moves.filter(m=>m.move||m.value).forEach(m=>children.push(bullet(safe(m.move||m.value))));
    children.push(blank());
  }
  if(partners.length){
    children.push(h2("Partnerships & Integrations"), blank());
    children.push(new Table({
      width:{size:9360,type:WidthType.DXA},
      columnWidths:[2160,1440,4320,1440],
      rows:[
        new TableRow({children:[hdrCell("Partner",2160),hdrCell("Type",1440),
                                hdrCell("Description",4320),hdrCell("Date",1440)]}),
        ...partners.map((p,i)=>{
          const shade = i%2===0?CGREY:CWHITE;
          return new TableRow({children:[
            boldCell(safe(p.partner||p.value),2160,i%2===0?"E8F0FC":CACCL),
            strCell(safe(p.type),1440,shade),
            strCell(safe(p.description),4320,shade),
            strCell(safe(p.date),1440,shade),
          ]});
        })
      ]
    }));
    children.push(blank());
  }
  children.push(pageBreak());
}

// ── 9. COMPETITORS & DIFFERENTIATORS ─────────────────────────────────
const competitors = D.competitors||[];
const diffs       = D.differentiators||[];
if(competitors.length||diffs.length){
  children.push(h1("Competitive Landscape"), blank());
  if(competitors.length){
    children.push(h2("Competitors"), blank());
    children.push(new Table({
      width:{size:9360,type:WidthType.DXA},
      columnWidths:[3240,6120],
      rows:[
        new TableRow({children:[hdrCell("Competitor",3240),hdrCell("Domain / Notes",6120)]}),
        ...competitors.map((c,i)=>{
          const shade = i%2===0?CGREY:CWHITE;
          const name  = safe(c.name||c.value);
          const dom   = safe(c.domain);
          return new TableRow({children:[
            boldCell(name,3240,i%2===0?"E8F0FC":CACCL),
            isUrl(dom)
              ? cell([para([link(dom,dom)])],6120,shade)
              : strCell(dom,6120,shade),
          ]});
        })
      ]
    }));
    children.push(blank());
  }
  if(diffs.length){
    children.push(h2("Differentiators"), blank());
    diffs.filter(d=>d.feature||d.value).forEach(d=>children.push(bullet(safe(d.feature||d.value))));
    children.push(blank());
  }
}

// ══════════════════════════════════════════════════════════════════════
// HEADER + FOOTER
// ══════════════════════════════════════════════════════════════════════
const headerPara = new Paragraph({
  border:{bottom:{style:BorderStyle.SINGLE,size:8,color:CACC,space:1}},
  spacing:{after:0},
  children:[
    bold(`${compName}`,{size:18,color:CDARK}),
    new TextRun({text:"\t",font:FONT}),
    run("Intelligence Report",{size:17,color:CMUTED}),
    new TextRun({text:"\t",font:FONT}),
    run(compDom,{size:17,color:CACC}),
  ],
  tabStops:[
    {type:TabStopType.CENTER,position:4680},
    {type:TabStopType.RIGHT, position:9360},
  ],
});

const footerPara = new Paragraph({
  border:{top:{style:BorderStyle.SINGLE,size:6,color:CLINE,space:1}},
  alignment:AlignmentType.CENTER,
  children:[
    run("Page ",{color:CMUTED,size:17}),
    new TextRun({children:[PageNumber.CURRENT],font:FONT,size:17,color:CMUTED}),
    run("  of  ",{color:CMUTED,size:17}),
    new TextRun({children:[PageNumber.TOTAL_PAGES],font:FONT,size:17,color:CMUTED}),
    run("   ·   Confidential",{color:CMUTED,size:17}),
  ],
});

// ══════════════════════════════════════════════════════════════════════
// ASSEMBLE DOC
// ══════════════════════════════════════════════════════════════════════
const doc = new Document({
  numbering,styles,
  sections:[{
    properties:{
      page:{
        size:{width:12240,height:15840},
        margin:{top:1080,right:1260,bottom:1080,left:1260},
      }
    },
    headers:{default:new Header({children:[headerPara]})},
    footers:{default:new Footer({children:[footerPara]})},
    children,
  }]
});

Packer.toBuffer(doc).then(buf=>{
  fs.writeFileSync(OUT_PATH,buf);
  console.log("docx ok:",OUT_PATH);
}).catch(e=>{console.error(e);process.exit(1);});
'''


def build_docx(d: ReportData, logo_bytes: bytes, out_path: str):
    """Serialise ReportData to JSON, inject into the JS template, run Node."""

    # Flatten data into a single dict for the JS side
    data_for_js = {
        "project_name":    d.project_name,
        "company_profile": {
            "company_name":                d.company_name,
            "domain":                      d.company_domain,
            "playstore_link":              d.playstore_link,
            "appstore_link":               d.appstore_link,
            "youtube_official_channel":    d.youtube_channel,
            "linkedin_company_page":       d.linkedin_page,
            "year_founded":                d.year_founded,
            "exact_hq_location":           d.hq_location,
            "locations_operating_in":      d.locations,
            "industry_and_segment":        d.industry,
            "available_platforms":         d.platforms,
            "employee_count":              d.employee_count,
            "funding_raised":              d.funding_raised,
            "funding_stage":               d.funding_stage,
            "annual_revenue":              d.annual_revenue,
            "key_positioning":             d.key_positioning,
            "revenue_model":               d.revenue_model,
            "pricing_tiers":               d.pricing_tiers,
            "target_customer_segments":    d.target_segments,
            "tech_stack_highlights":       d.tech_stack,
            "market_sentiment":            d.market_sentiment,
            "c-suite_officer":             d.csuite,
            "names_of_founders":           d.founders,
        },
        "play_store":       d.ps,
        "app_store":        d.app,
        "reviews_list":     d.ps.get("reviews_list", []),
        "transcript_meta":  {
            "source_file":   d.transcript.get("source_file"),
            "total_signals": d.transcript.get("total_signals"),
            "classifier":    d.transcript.get("classifier"),
            "meeting_type":  d.transcript.get("meeting_type"),
        },
        "signals":          d.transcript.get("signals", []),
        "reddit_posts":     d.reddit_posts,
        "youtube_videos":   d.youtube_videos,
        "problems":         d.problems,
        "total_problems":   d.total_problems,
        "top_categories":   d.top_categories,
        "high_severity_count": d.high_severity_count,
        "sources_used":     "",
        "insights":         d.insights,
        "total_insights":   d.total_insights,
        "critical_count":   d.critical_count,
        "dominant_theme":   d.dominant_theme,
        "strategic_risk":   d.strategic_risk,
        "biggest_opp":      d.biggest_opp,
        "briefs":           d.briefs,
        "total_briefs":     d.total_briefs,
        "sprint_focus":     d.sprint_focus,
        "milestones":       d.milestones,
        "strategic_moves":  d.strategic_moves,
        "partnerships":     d.partnerships,
        "competitors":      d.competitors,
        "differentiators":  d.differentiators,
    }

    with tempfile.TemporaryDirectory() as tmpdir:
        data_path = os.path.join(tmpdir, "data.json")
        with open(data_path, "w", encoding="utf-8") as f:
            json.dump(data_for_js, f, ensure_ascii=False, indent=2)

        logo_path = ""
        if logo_bytes:
            logo_path = os.path.join(tmpdir, "logo.png")
            with open(logo_path, "wb") as f:
                f.write(logo_bytes)

        js_code = (DOCX_JS_TEMPLATE
                   .replace("__DATA_JSON__", data_path)
                   .replace("__LOGO_FILE__",  logo_path)
                   .replace("__OUT_PATH__",   out_path))

        js_path = os.path.join(tmpdir, "gen.js")
        with open(js_path, "w", encoding="utf-8") as f:
            f.write(js_code)

        result = subprocess.run(["node", js_path], capture_output=True, text=True)
        if result.returncode != 0:
            print("  ✗ Node error:", result.stderr[:600])
            raise RuntimeError("docx generation failed")
        print(f"  ✓ DOCX   →  {out_path}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main(json_path: str):
    print(f"\n{'='*60}")
    print(f"  Report Generator")
    print(f"  Input : {json_path}")
    print(f"{'='*60}\n")

    # 1. Parse
    print("  Parsing JSON …")
    d = ReportData(json_path)
    print(f"  Company  : {d.company_name}")
    print(f"  Domain   : {clean_domain(d.company_domain)}")

    # 2. Fetch logo
    print("  Fetching logo …")
    logo_bytes, _ = fetch_logo_bytes(d.company_domain, d.company_name)
    print(f"  Logo     : {len(logo_bytes)} bytes")

    # 3. Create output directory
    out_dir = Path("output") / slug(d.company_name)
    out_dir.mkdir(parents=True, exist_ok=True)
    base = slug(d.company_name)

    # 4. Build Excel
    print("  Building Excel …")
    ExcelBuilder().build(d, logo_bytes, str(out_dir / f"{base}_report.xlsx"))

    # 5. Build DOCX
    print("  Building DOCX …")
    build_docx(d, logo_bytes, str(out_dir / f"{base}_report.docx"))

    print(f"\n  Done!  Output folder: {out_dir.resolve()}\n")


if __name__ == "__main__":
    json_file = sys.argv[1] if len(sys.argv) > 1 else "db_document.json"
    main(json_file)